from flask import Flask, request, jsonify, render_template
import cohere  # Ensure you have the Cohere API library installed
import os
import openpyxl
from datetime import datetime
import uuid

app = Flask(__name__)

# Initialize the Cohere client
api_key = '3v1gSXcJA8TXjPUHe0kPcCyzEMY7Qo5f52M82SuN'
co = cohere.Client(api_key)

def generate_unique_filename(base_name):
    # Generate a 2-digit random number
    random_number = str(uuid.uuid4().int)[:2]  # Take the first 2 digits from UUID
    # Get the current timestamp
    timestamp = datetime.now().strftime('%Y%m%d%H%M')
    # Combine base name, random number, and timestamp to form a unique filename
    unique_filename = f"{base_name}_{random_number}_{timestamp}"
    return unique_filename

# Define a function to read prompt from file and format it with user data
def load_and_format_prompt(prompt_name, data):
    file_path = os.path.join('prompts', f'{prompt_name}_prompt.txt')
    excel_path = "userData.xlsx"

    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Prompt file for {prompt_name} not found.")
    
    # Read the prompt template
    with open(file_path, 'r') as file:
        prompt_template = file.read()
    
    # Ensure that all keys in data have corresponding placeholders in the template
    try:
        prompt = prompt_template.format(**data)
    except KeyError as e:
        raise KeyError(f"Missing key in template: {e}")

    # Get current date and time
    current_date = datetime.now().date()
    current_time = datetime.now().time()

    # Load or create the Excel workbook
    if os.path.exists(excel_path):
        workbook = openpyxl.load_workbook(excel_path)
    else:
        workbook = openpyxl.Workbook()
        # Remove the default sheet created
        if 'Sheet1' in workbook.sheetnames:
            default_sheet = workbook['Sheet1']
            workbook.remove(default_sheet)
    
    # Select or create a sheet based on the prompt name
    if prompt_name in workbook.sheetnames:
        sheet = workbook[prompt_name]
    else:
        sheet = workbook.create_sheet(title=prompt_name)

    # Write the data to the selected sheet according to column headers
    if sheet.max_row == 1 and sheet.max_column == 1:  # Check if the sheet is new and empty
        # Add headers
        headers = list(data.keys()) + ['Date', 'Time']
        sheet.append(headers)
    else:
        headers = [cell.value for cell in sheet[1]]
    # Prepare the row data
    row_data = [data.get(header, '') for header in headers if header in data]
    row_data += [current_date, current_time]

    # Append the row data to the sheet
    sheet.append(row_data)

    # Save the workbook
    workbook.save(excel_path)
    return prompt

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/business-growth', methods=['GET', 'POST'])
def business_growth():
    if request.method == 'POST':
        data = request.json
        prompt = load_and_format_prompt('business_growth', data)
        response = co.generate(prompt=prompt, model='command-r-plus', temperature=0.5)
        strategy = response.generations[0].text
        return jsonify({'strategy': strategy})
    return render_template('business_growth.html')

@app.route('/lead-generation', methods=['GET', 'POST'])
def lead_generation():
    if request.method == 'POST':
        data = request.json
        prompt = load_and_format_prompt('lead_gen', data)
        response = co.generate(prompt=prompt, model='command-r-plus', temperature=0.5)
        strategy = response.generations[0].text
        return jsonify({'strategy': strategy})
    return render_template('lead_generation.html')

@app.route('/funding-pitch', methods=['GET', 'POST'])
def funding_pitch():
    if request.method == 'POST':
        data = request.json
        prompt = load_and_format_prompt('funding_pitch', data)
        response = co.generate(prompt=prompt, model='command-r-plus', temperature=0.5)
        strategy = response.generations[0].text
        return jsonify({'strategy': strategy})
    return render_template('funding_pitch.html')

@app.route('/social-media-strategy', methods=['GET', 'POST'])
def social_media_strategy():
    if request.method == 'POST':
        data = request.json
        prompt = load_and_format_prompt('social_media_strategy', data)
        response = co.generate(prompt=prompt, model='command-r-plus', temperature=0.5)
        strategy = response.generations[0].text
        return jsonify({'strategy': strategy})
    return render_template('social_media_strategy.html')

@app.route('/business-queries', methods=['GET', 'POST'])
def business_queries():
    if request.method == 'POST':
        data = request.json
        prompt = load_and_format_prompt('business_queries', data)
        response = co.generate(prompt=prompt, model='command-r-plus', temperature=0.5)
        strategy = response.generations[0].text
        return jsonify({'strategy': strategy})
    return render_template('business_queries.html')

@app.route('/linkedin-strategy', methods=['GET', 'POST'])
def linkedin_strategy():
    if request.method == 'POST':
        data = request.json
        prompt = load_and_format_prompt('linkedin_strategy', data)
        response = co.generate(prompt=prompt, model='command-r-plus', temperature=0.5)
        strategy = response.generations[0].text
        return jsonify({'strategy': strategy})
    return render_template('linkedin_strategy.html')

if __name__ == '__main__':
    app.run(debug=True)
